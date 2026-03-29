from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import json
import os
import qrcode
import io
import base64

from xhtml2pdf import pisa
from .models import Department, Student, Subject, Mark, Year, Semester
from .utils import calculate_grade


# ──────────────────────────────────────────────
# PDF Download
# ──────────────────────────────────────────────
def download_pdf(request, student_id):
    student = Student.objects.get(id=student_id)
    marks = Mark.objects.filter(student=student)

    template = get_template('results/pdf.html')
    html = template.render({'student': student, 'marks': marks})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="marksheet.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response


# ──────────────────────────────────────────────
# Admin Login
# ──────────────────────────────────────────────
def admin_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        if username == "admin" and password == "admin123":
            request.session['admin'] = True
            return redirect('admin_dashboard')
        else:
            return render(request, 'results/admin_login.html', {'error': 'Invalid credentials'})
    return render(request, 'results/admin_login.html')


# ──────────────────────────────────────────────
# Admin Dashboard
# ──────────────────────────────────────────────
import json
def get_best_marks(marks):
    def is_pass(m): return m.external >= 30
    cleared = {}
    failed = {}
    sorted_marks = sorted(marks, key=lambda x: (x.semester, x.id))
    for m in sorted_marks:
        sid = m.subject_id
        if is_pass(m):
            cleared[sid] = m
            failed.pop(sid, None)
        else:
            if sid not in cleared:
                failed[sid] = m
    return list(cleared.values()) + list(failed.values())

def admin_dashboard(request):
    dept_id = request.GET.get('department')
    year_id = request.GET.get('year')
    sem_id = request.GET.get('semester')
    sort_by = request.GET.get('sort', 'roll_no')
    order = request.GET.get('order', 'asc')

    students_query = Student.objects.all()
    
    if dept_id:
        students_query = students_query.filter(department_id=dept_id)
    if year_id:
        students_query = students_query.filter(year_id=year_id)
    if sem_id:
        students_query = students_query.filter(semester__name=sem_id)

    students = list(students_query)
    
    # Calculate marks and percentages for sorting and display
    for s in students:
        marks = list(s.mark_set.all())
        best_marks = get_best_marks(marks)
        # Always compute total live as internal+external (same as marksheet view)
        # so that stale or None stored totals don't affect the dashboard
        computed_totals = [(m.internal + m.external) for m in best_marks]
        s.total_mark_val = sum(computed_totals)
        n = len(computed_totals)
        s.percentage_val = round((s.total_mark_val / (n * 100)) * 100, 2) if n > 0 else 0

    # Sort
    reverse_sort = (order == 'desc')
    if sort_by == 'mark':
        students.sort(key=lambda x: x.total_mark_val, reverse=reverse_sort)
    elif sort_by == 'percentage':
        students.sort(key=lambda x: x.percentage_val, reverse=reverse_sort)
    else:
        # default roll_no
        students.sort(key=lambda x: x.roll_no, reverse=reverse_sort)

    departments = Department.objects.all()
    years = Year.objects.all()
    years_data = [{"id": y.id, "label": str(y), "dept_id": y.department_id} for y in years]

    context = {
        "students": students,
        "students_count": Student.objects.count(),
        "subjects_count": Subject.objects.count(),
        "departments_count": Department.objects.count(),
        "departments": departments,
        "years": years,
        "years_json": json.dumps(years_data),
        "sel_dept": int(dept_id) if dept_id else "",
        "sel_year": int(year_id) if year_id else "",
        "sel_sem": int(sem_id) if sem_id else "",
        "sort_by": sort_by,
        "order": order,
    }
    return render(request, "results/admin_dashboard.html", context)


# ──────────────────────────────────────────────
# Student Login
# ──────────────────────────────────────────────
def student_login(request):
    if request.method == "POST":
        roll_no = request.POST['roll_no']
        dob = request.POST['dob']
        try:
            student = Student.objects.get(roll_no=roll_no, dob=dob)
            return redirect('marksheet', student_id=student.id)
        except Student.DoesNotExist:
            return render(request, 'results/login.html', {'error': 'Invalid Roll Number or DOB'})
    return render(request, 'results/login.html')


# ──────────────────────────────────────────────
# Marksheet
# ──────────────────────────────────────────────
def marksheet(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    # QR Code – generate as inline base64 data URI (no file/media serving needed)
    marksheet_url = request.build_absolute_uri()
    qr_img = qrcode.make(marksheet_url)
    buffer = io.BytesIO()
    qr_img.save(buffer, format="PNG")
    qr_b64 = base64.b64encode(buffer.getvalue()).decode()
    qr_url = f"data:image/png;base64,{qr_b64}"

    # ── Fetch all marks ordered by semester ──────────────────────
    all_marks = Mark.objects.filter(student=student).select_related('subject').order_by('semester')

    # ── Grade helper ──────────────────────────────────────────────
    def get_grade(total):
        if total >= 90: return "O", 10
        if total >= 80: return "A+", 9
        if total >= 70: return "A", 8
        if total >= 60: return "B+", 7
        if total >= 50: return "B", 6
        if total >= 40: return "C", 5
        return "F", 0

    # ── Helper: determine pass/fail/absent from external mark ─────
    def is_pass(mark):   return mark.external >= 30
    def is_arrear(mark): return 1 <= mark.external <= 29
    def is_absent(mark): return mark.external == 0
    def is_failed(mark): return mark.external <= 29  # absent or arrear

    # ── Annotate every mark with grade ────────────────────────────
    for mark in all_marks:
        mark.total_calc = mark.internal + mark.external
        if is_pass(mark):
            mark.grade, mark.grade_point = get_grade(mark.total_calc)
        else:
            mark.grade, mark.grade_point = "-", 0
        mark.points = mark.grade_point * mark.subject.credits

    # ── Group marks by semester ────────────────────────────────────
    from collections import defaultdict
    sem_groups = defaultdict(list)
    for mark in all_marks:
        sem_groups[mark.semester].append(mark)

    # ── Arrear tracking (based on external marks) ─────────────────
    cleared_subjects = {}  # subject_id → mark that cleared it
    failed_subjects  = {}  # subject_id → latest failing mark

    sorted_sems = sorted(sem_groups.keys())
    for sem in sorted_sems:
        for mark in sem_groups[sem]:
            subj_id = mark.subject_id
            if is_pass(mark):
                cleared_subjects[subj_id] = mark
                failed_subjects.pop(subj_id, None)
            else:  # absent or arrear
                if subj_id not in cleared_subjects:
                    failed_subjects[subj_id] = mark

    # ── Build per-semester display data ───────────────────────────
    semester_data = []
    running_arrears = set()  # subject_ids with pending arrears entering each semester

    for sem in sorted_sems:
        marks_in_sem = sem_groups[sem]
        sem_total_marks  = 0
        sem_total_points = 0
        sem_total_credits = 0

        # Separate regular subjects vs arrear re-attempts
        regular_marks       = []
        arrear_attempt_marks = []
        for m in marks_in_sem:
            if m.subject_id in running_arrears:
                arrear_attempt_marks.append(m)
            else:
                regular_marks.append(m)

        for mark in marks_in_sem:
            sem_total_marks += mark.total_calc
            if is_pass(mark):
                sem_total_points  += mark.points
                sem_total_credits += mark.subject.credits

        n = len(marks_in_sem)
        sem_percentage = round((sem_total_marks / (n * 100)) * 100, 2) if n else 0
        sem_gpa = round(sem_total_points / sem_total_credits, 2) if sem_total_credits else 0

        # Update running arrears AFTER this semester
        newly_failed    = {m.subject_id for m in marks_in_sem if is_failed(m)}
        cleared_this_sem = {m.subject_id for m in marks_in_sem if is_pass(m)}
        running_arrears = (running_arrears | newly_failed) - cleared_this_sem

        semester_data.append({
            "semester":       sem,
            "marks":          marks_in_sem,
            "regular_marks":  regular_marks,
            "arrear_attempts": arrear_attempt_marks,
            "total_marks":    sem_total_marks,
            "percentage":     sem_percentage,
            "gpa":            sem_gpa,
            "arrears_after":  len(running_arrears),
            "failed_count":   len(newly_failed),
        })

    # ── Still-pending arrears ─────────────────────────────────────
    pending_arrears = list(failed_subjects.values())

    # ── Overall totals ────────────────────────────────────────────
    best_marks    = list(cleared_subjects.values()) + list(failed_subjects.values())
    total_marks   = sum(m.total_calc for m in best_marks)
    n_all         = len(best_marks)
    total_points  = sum(m.points for m in best_marks if is_pass(m))
    total_credits = sum(m.subject.credits for m in best_marks if is_pass(m))
    percentage    = round((total_marks / (n_all * 100)) * 100, 2) if n_all else 0
    gpa           = round(total_points / total_credits, 2) if total_credits else 0

    context = {
        "student": student,
        "all_marks": all_marks,
        "semester_data": semester_data,
        "pending_arrears": pending_arrears,
        "total_arrears": len(pending_arrears),
        "total_marks": total_marks,
        "percentage": percentage,
        "gpa": gpa,
        "month_year": "Nov 2023",
        "qr_code_url": qr_url,
    }
    return render(request, "results/marksheet.html", context)



# ──────────────────────────────────────────────
# Manual Add Students (Handsontable sheet)
# ──────────────────────────────────────────────
def _get_or_create_semester(year, semester_int):
    """Get or create a Semester record for a given year and semester number."""
    semester_obj, _ = Semester.objects.get_or_create(
        year=year,
        name=str(semester_int)
    )
    return semester_obj


def add_students_sheet(request, department_id, year_id, semester_id):
    department = Department.objects.get(id=department_id)
    year = Year.objects.get(id=year_id)
    semester = int(semester_id)

    students = Student.objects.filter(department=department, year=year)
    subjects = Subject.objects.filter(department=department, semester=semester)

    students_list = []
    for s in students:
        students_list.append([s.roll_no, s.name, s.dob.strftime('%Y-%m-%d') if s.dob else ""])

    context = {
        "students_json": json.dumps(students_list),
        "year": year,
        "semester_id": semester,
        "department_id": department_id,
        "year_id": year_id,
    }
    return render(request, "results/add_students_sheet.html", context)


@csrf_exempt
def save_students_api(request):
    """Save students from Handsontable manual entry."""
    if request.method == "POST":
        data = json.loads(request.body)
        rows = data.get("data")
        department_id = data.get("department_id")
        year_id = data.get("year_id")
        semester_id = data.get("semester_id")

        department = Department.objects.get(id=int(department_id))
        year = Year.objects.get(id=int(year_id))
        semester_obj = _get_or_create_semester(year, int(semester_id))

        # Collect non-empty roll numbers
        roll_numbers = [row[0] for row in rows if row[0]]

        # Check duplicates within the submitted sheet
        if len(roll_numbers) != len(set(roll_numbers)):
            return JsonResponse({"error": "Duplicate roll numbers found in sheet"}, status=400)

        created_count = 0
        skipped = []
        for row in rows:
            roll_no = row[0]
            if not roll_no:
                continue  # skip blank rows
            name = row[1]
            dob = row[2]
            try:
                _, created = Student.objects.update_or_create(
                    roll_no=roll_no,
                    defaults={
                        "name": name, "dob": dob,
                        "department": department, "year": year,
                        "semester": semester_obj
                    }
                )
                if created:
                    created_count += 1
                else:
                    skipped.append(roll_no)
            except Exception as e:
                return JsonResponse({"error": f"Error saving {roll_no}: {str(e)}"}, status=400)

        msg = f"Saved {created_count} student(s)."
        if skipped:
            msg += f" Skipped (already exist): {', '.join(skipped)}"
        return JsonResponse({"status": "success", "message": msg})

    return JsonResponse({"error": "Invalid method"}, status=405)


# ──────────────────────────────────────────────
# Manual Add Marks (Handsontable sheet)
# ──────────────────────────────────────────────
def add_marks_sheet(request, department_id, year_id, semester_id):
    from .models import Department, Student, Subject, Mark, Semester
    department = Department.objects.get(id=department_id)
    semester_int = int(semester_id)

    # Fetch all students enrolled in this year and department batch
    students = list(Student.objects.filter(department=department, year_id=year_id))

    # Regular subjects for this semester
    subjects = list(Subject.objects.filter(department=department, semester=semester_int))

    # ── Compute pending arrear subjects per student ───────────────────
    # A subject is a pending arrear if:
    #   - The student has a mark for it in a semester < current semester
    #   - The latest mark for that subject has external == 0 or external <= 29 (not cleared)
    #   - The student has NOT already cleared it (external >= 30) in any attempt
    def get_pending_arrears(student):
        """Return list of (subject, original_semester) tuples still pending for student."""
        all_marks = (
            Mark.objects.filter(student=student, semester__lt=semester_int)
            .select_related('subject')
            .order_by('semester')
        )
        # Group by subject → track cleared/failed state
        subject_state = {}  # subject_id → {'subject': ..., 'orig_sem': ..., 'cleared': bool}
        for m in all_marks:
            sid = m.subject_id
            is_cleared = m.external >= 30
            if sid not in subject_state:
                subject_state[sid] = {'subject': m.subject, 'orig_sem': m.semester, 'cleared': is_cleared}
            else:
                # Later attempts override: once cleared, always cleared
                if is_cleared:
                    subject_state[sid]['cleared'] = True
                elif not subject_state[sid]['cleared']:
                    subject_state[sid]['cleared'] = False

        pending = []
        for sid, state in subject_state.items():
            if not state['cleared']:
                pending.append((state['subject'], state['orig_sem']))
        return pending

    # Build per-student arrear info
    student_arrears = {}  # student.roll_no → list of (subject, orig_sem)
    for student in students:
        student_arrears[student.roll_no] = get_pending_arrears(student)

    # Union of all arrear subjects across all students (for column headers)
    # Key: (subject_id, orig_sem) → (subject, orig_sem)
    all_arrear_cols = {}
    for roll_no, arrears in student_arrears.items():
        for subj, orig_sem in arrears:
            key = (subj.id, orig_sem)
            if key not in all_arrear_cols:
                all_arrear_cols[key] = (subj, orig_sem)

    # Sort arrear columns by orig_sem then subject id for consistent ordering
    sorted_arrear_cols = sorted(all_arrear_cols.values(), key=lambda x: (x[1], x[0].id))

    # ── Build headers ─────────────────────────────────────────────────
    headers = ["Reg No", "Name"]
    for subject in subjects:
        headers.append(f"{subject.name} (Int)")
        headers.append(f"{subject.name} (Ext)")
    for subj, orig_sem in sorted_arrear_cols:
        headers.append(f"{subj.name} [Arrear Sem {orig_sem}] (Int)")
        headers.append(f"{subj.name} [Arrear Sem {orig_sem}] (Ext)")

    # ── Build data rows ───────────────────────────────────────────────
    data = []
    for student in students:
        row = [student.roll_no, student.name]

        # Regular subject marks
        for subject in subjects:
            try:
                mark = Mark.objects.get(student=student, subject=subject, semester=semester_int)
                row += [mark.internal, mark.external]
            except Mark.DoesNotExist:
                row += [None, None]

        # Arrear subject marks for this student (blank if this student has no arrear in that col)
        student_pending_keys = {(s.id, sem) for s, sem in student_arrears.get(student.roll_no, [])}
        for subj, orig_sem in sorted_arrear_cols:
            key = (subj.id, orig_sem)
            if key in student_pending_keys:
                # Check if re-attempt mark already saved for this current semester
                try:
                    mark = Mark.objects.get(student=student, subject=subj, semester=semester_int)
                    row += [mark.internal, mark.external]
                except Mark.DoesNotExist:
                    row += [None, None]
            else:
                # This student doesn't have this arrear – show N/A as read-only placeholder
                row += ["N/A", "N/A"]

        data.append(row)

    # ── Column read-only config for Handsontable ──────────────────────
    # First 2 cols (Reg No, Name) are read-only always.
    # Arrear columns where student has "N/A" are also read-only (handled in JS).
    num_regular_cols = 2 + len(subjects) * 2

    # Pass arrear meta to template for JS awareness
    arrear_col_meta = []
    for subj, orig_sem in sorted_arrear_cols:
        arrear_col_meta.append({"subject_id": subj.id, "orig_sem": orig_sem, "subject_name": subj.name})

    context = {
        "headers_json": json.dumps(headers),
        "data_json": json.dumps(data),
        "semester_id": semester_id,
        "department_id": department_id,
        "year_id": year_id,
        "num_regular_cols": num_regular_cols,
        "arrear_col_meta_json": json.dumps(arrear_col_meta),
        "has_arrears": len(sorted_arrear_cols) > 0,
    }
    return render(request, "results/add_marks_sheet.html", context)


@csrf_exempt
def save_marks_api(request):
    """Save marks from Handsontable entry. Handles both regular and arrear columns."""
    print("SAVE MARKS API HIT")

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=400)

    try:
        data = json.loads(request.body)
        rows = data.get("data", [])
        department_id = data.get("department_id")
        year_id = data.get("year_id")
        semester_id = data.get("semester_id")

        if not all([rows, department_id, year_id, semester_id]):
            return JsonResponse({"error": "Missing required data"}, status=400)

        department = Department.objects.get(id=department_id)
        year = Year.objects.get(id=year_id)
        semester_int = int(semester_id)

        # Regular subjects for this semester (same order as in add_marks_sheet)
        regular_subjects = list(Subject.objects.filter(department=department, semester=semester_int))

        # Helper: replicate get_pending_arrears logic for a student
        def get_pending_arrears(student):
            all_marks = (
                Mark.objects.filter(student=student, semester__lt=semester_int)
                .select_related('subject')
                .order_by('semester')
            )
            subject_state = {}
            for m in all_marks:
                sid = m.subject_id
                is_cleared = m.external >= 30
                if sid not in subject_state:
                    subject_state[sid] = {'subject': m.subject, 'orig_sem': m.semester, 'cleared': is_cleared}
                else:
                    if is_cleared:
                        subject_state[sid]['cleared'] = True
                    elif not subject_state[sid]['cleared']:
                        subject_state[sid]['cleared'] = False
            pending = []
            for sid, state in subject_state.items():
                if not state['cleared']:
                    pending.append((state['subject'], state['orig_sem']))
            return pending

        # Helper: save a single mark cell pair
        def save_mark_cell(student, subject, sem, internal_raw, external_raw):
            try:
                internal = int(internal_raw) if internal_raw not in [None, "", "N/A"] else None
            except Exception:
                internal = None
            try:
                external = int(external_raw) if external_raw not in [None, "", "N/A"] else None
            except Exception:
                external = None

            if internal is None and external is None:
                return  # N/A placeholder – skip

            internal = internal if internal is not None else 0
            external = external if external is not None else 0

            Mark.objects.update_or_create(
                student=student, subject=subject, semester=sem,
                defaults={"internal": internal, "external": external}
            )

        # Build global arrear column order (same as add_marks_sheet)
        # We need to collect all students first to build the union of arrear cols
        all_students = {}
        for row in rows:
            if not row or not row[0]:
                continue
            roll_no = str(row[0])
            try:
                student = Student.objects.get(roll_no=roll_no, department=department, year=year)
                all_students[roll_no] = student
            except Student.DoesNotExist:
                continue

        student_arrears = {roll: get_pending_arrears(s) for roll, s in all_students.items()}

        all_arrear_cols = {}
        for roll_no, arrears in student_arrears.items():
            for subj, orig_sem in arrears:
                key = (subj.id, orig_sem)
                if key not in all_arrear_cols:
                    all_arrear_cols[key] = (subj, orig_sem)

        sorted_arrear_cols = sorted(all_arrear_cols.values(), key=lambda x: (x[1], x[0].id))

        # Process each data row
        for row in rows:
            if not row or not row[0]:
                continue
            roll_no = str(row[0])
            student = all_students.get(roll_no)
            if not student:
                continue

            col_index = 2

            # Save regular subject marks
            for subject in regular_subjects:
                internal = row[col_index] if col_index < len(row) else None
                external = row[col_index + 1] if col_index + 1 < len(row) else None
                save_mark_cell(student, subject, semester_int, internal, external)
                col_index += 2

            # Save arrear re-attempt marks (stored under current semester)
            student_pending_keys = {(s.id, sem) for s, sem in student_arrears.get(roll_no, [])}
            for subj, orig_sem in sorted_arrear_cols:
                internal = row[col_index] if col_index < len(row) else None
                external = row[col_index + 1] if col_index + 1 < len(row) else None
                key = (subj.id, orig_sem)
                if key in student_pending_keys:
                    # Save arrear re-attempt under current semester
                    save_mark_cell(student, subj, semester_int, internal, external)
                # If N/A (student has no arrear for this col), skip silently
                col_index += 2

        return JsonResponse({"status": "success"})

    except Exception as e:
        print("ERROR:", str(e))
        return JsonResponse({"error": str(e)}, status=500)



# ──────────────────────────────────────────────
# Select screens
# ──────────────────────────────────────────────

def get_semesters_for_year(request, year_id):
    """
    Return semesters available for a given year.
    Logic: year position within its department determines semester range.
      Position 1 (1st year) → [1, 2]
      Position 2 (2nd year) → [1, 2, 3, 4]
      Position 3 (3rd year) → [1, 2, 3, 4, 5, 6]
    """
    try:
        year = Year.objects.get(id=year_id)
        # Get ordered years for this department
        dept_years = list(Year.objects.filter(department=year.department).order_by('id'))
        position = next((i + 1 for i, y in enumerate(dept_years) if y.id == year.id), 1)
        max_sem = min(position * 2, 6)
        semesters = list(range(1, max_sem + 1))
        return JsonResponse({"semesters": semesters, "position": position})
    except Year.DoesNotExist:
        return JsonResponse({"semesters": [1, 2]})


def delete_student(request, student_id):
    """Delete a student and all their marks."""
    student = get_object_or_404(Student, id=student_id)
    if request.method == "POST":
        student.delete()
    return redirect("admin_dashboard")


def select_marks(request):
    departments = Department.objects.all()
    years = Year.objects.all()

    if request.method == "POST":
        department_id = request.POST.get("department")
        year_id = request.POST.get("year")
        semester_id = request.POST.get("semester")
        return redirect("add_marks_sheet", department_id=department_id, year_id=year_id, semester_id=semester_id)

    # Build years with department info for JS filtering
    years_data = [{"id": y.id, "label": str(y), "dept_id": y.department_id} for y in years]

    return render(request, "results/select_marks.html", {
        "departments": departments,
        "years": years,
        "years_json": json.dumps(years_data),
    })


def select_students(request):
    departments = Department.objects.all()
    years = Year.objects.all()

    if request.method == "POST":
        department_id = request.POST.get("department")
        year_id = request.POST.get("year")
        semester_id = request.POST.get("semester")
        return redirect("add_students_sheet", department_id=department_id, year_id=year_id, semester_id=semester_id)

    years_data = [{"id": y.id, "label": str(y), "dept_id": y.department_id} for y in years]

    return render(request, "results/select_students.html", {
        "departments": departments,
        "years": years,
        "years_json": json.dumps(years_data),
    })


# ──────────────────────────────────────────────
# Excel Import – shared helper
# ──────────────────────────────────────────────
def _build_years_json():
    """Build the years JSON payload needed by the dept→year JS filter on import pages."""
    years = Year.objects.all()
    data = [{"id": y.id, "label": str(y), "dept_id": y.department_id} for y in years]
    return json.dumps(data)


# ──────────────────────────────────────────────
# Excel Import – Students
# ──────────────────────────────────────────────
def import_students_page(request):
    departments = Department.objects.all()
    years = Year.objects.all()
    years_data = [{"id": y.id, "label": str(y), "dept_id": y.department_id} for y in years]
    return render(request, "results/import_students.html", {
        "departments": departments,
        "years": years,
        "years_json": json.dumps(years_data),
        "semesters": [1, 2, 3, 4, 5, 6],
    })


def import_students_excel(request):
    """Handle .xlsx upload, save to disk, and bulk-create students."""
    if request.method != "POST":
        return redirect("import_students_page")

    try:
        import openpyxl
    except ImportError:
        return render(request, "results/import_students.html", {
            "error": "openpyxl is not installed. Run: pip install openpyxl",
            "departments": Department.objects.all(),
            "years": Year.objects.all(),
            "years_json": _build_years_json(),
        })

    department_id = request.POST.get("department")
    year_id = request.POST.get("year")
    semester_id = request.POST.get("semester")
    excel_file = request.FILES.get("excel_file")

    def _ctx_base():
        return {
            "departments": Department.objects.all(),
            "years": Year.objects.all(),
            "years_json": _build_years_json(),
            "semesters": [1, 2, 3, 4, 5, 6],
        }

    errors = []
    if not department_id:
        errors.append("Please select a department.")
    if not year_id:
        errors.append("Please select a year.")
    if not semester_id:
        errors.append("Please select a semester.")
    if not excel_file:
        errors.append("Please upload an Excel file.")
    elif not excel_file.name.endswith(".xlsx"):
        errors.append("Only .xlsx files are supported.")

    if errors:
        ctx = _ctx_base()
        ctx["errors"] = errors
        return render(request, "results/import_students.html", ctx)

    try:
        department = Department.objects.get(id=department_id)
        year = Year.objects.get(id=year_id)
    except (Department.DoesNotExist, Year.DoesNotExist):
        ctx = _ctx_base()
        ctx["errors"] = ["Invalid department or year selected."]
        return render(request, "results/import_students.html", ctx)

    # ── Save the uploaded file to disk ─────────────────────────────────────
    import uuid
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile

    safe_name = f"{department_id}_{year_id}_{semester_id}_{uuid.uuid4().hex[:8]}_{excel_file.name}"
    save_path = os.path.join("uploads", "students", safe_name)
    saved_file_path = default_storage.save(save_path, ContentFile(excel_file.read()))
    # Reset pointer so openpyxl can read from the stored file
    full_saved_path = os.path.join(settings.MEDIA_ROOT, saved_file_path)

    try:
        wb = openpyxl.load_workbook(full_saved_path)
        ws = wb.active

        # ── Validate header row – students file must have exactly 3 columns ──
        header_row = [
            str(c).strip()
            for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if c is not None and str(c).strip() != ""
        ]
        if len(header_row) > 3:
            wrong_cols = ", ".join(header_row[:5]) + ("..." if len(header_row) > 5 else "")
            ctx = _ctx_base()
            ctx["errors"] = [
                f"❌ Wrong file uploaded! This looks like a Marks Excel file because it has "
                f"{len(header_row)} columns (Found: {wrong_cols}). "
                "The Students import file must have exactly 3 columns: Register No, Name, DOB. "
                "Please upload the correct Students Excel file or use the 'Download Sample' button."
            ]
            return render(request, "results/import_students.html", ctx)

        semester_obj = _get_or_create_semester(year, int(semester_id))

        created_count = 0
        skipped = []
        row_errors = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            roll_no = str(row[0]).strip()
            name = str(row[1]).strip() if row[1] else ""
            dob_raw = row[2]

            if not roll_no or not name or not dob_raw:
                row_errors.append(f"Row {i}: Missing data (roll_no, name, or DOB).")
                continue

            if hasattr(dob_raw, 'strftime'):
                dob = dob_raw.strftime('%Y-%m-%d')
            else:
                dob = str(dob_raw).strip()

            try:
                _, created = Student.objects.update_or_create(
                    roll_no=roll_no,
                    defaults={
                        "name": name, "dob": dob,
                        "department": department, "year": year,
                        "semester": semester_obj,
                    }
                )
                if created:
                    created_count += 1
                else:
                    skipped.append(roll_no)
            except Exception as e:
                row_errors.append(f"Row {i} ({roll_no}): {str(e)}")

        ctx = _ctx_base()
        ctx.update({
            "success": True,
            "created_count": created_count,
            "skipped": skipped,
            "row_errors": row_errors,
            "saved_file": safe_name,
        })
        return render(request, "results/import_students.html", ctx)

    except Exception as e:
        ctx = _ctx_base()
        ctx["errors"] = [f"Failed to read Excel file: {str(e)}"]
        return render(request, "results/import_students.html", ctx)


def download_student_sample_excel(request):
    """Return a sample .xlsx for the Students import format."""
    try:
        import openpyxl
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Register No", "Name", "DOB (YYYY-MM-DD)"])
    ws.append(["21CS001", "John Doe", "2003-05-10"])
    ws.append(["21CS002", "Jane Smith", "2003-08-22"])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="students_sample.xlsx"'
    wb.save(response)
    return response


# ──────────────────────────────────────────────
# Excel Import – Marks
# ──────────────────────────────────────────────
def import_marks_page(request):
    departments = Department.objects.all()
    years = Year.objects.all()
    years_data = [{"id": y.id, "label": str(y), "dept_id": y.department_id} for y in years]
    return render(request, "results/import_marks.html", {
        "departments": departments,
        "years": years,
        "years_json": json.dumps(years_data),
        "semesters": [1, 2, 3, 4, 5, 6],
    })


def import_marks_excel(request):
    """Handle .xlsx upload, save to disk, and bulk-upsert marks with cross-validation."""
    if request.method != "POST":
        return redirect("import_marks_page")

    try:
        import openpyxl
    except ImportError:
        return render(request, "results/import_marks.html", {
            "error": "openpyxl is not installed.",
            "departments": Department.objects.all(),
            "years": Year.objects.all(),
            "semesters": [1, 2, 3, 4, 5, 6],
            "years_json": _build_years_json(),
        })

    department_id = request.POST.get("department")
    year_id = request.POST.get("year")
    semester_id = request.POST.get("semester")
    excel_file = request.FILES.get("excel_file")

    ctx = {
        "departments": Department.objects.all(),
        "years": Year.objects.all(),
        "semesters": [1, 2, 3, 4, 5, 6],
        "years_json": _build_years_json(),
    }

    errors = []
    if not department_id:
        errors.append("Please select a department.")
    if not year_id:
        errors.append("Please select a year.")
    if not semester_id:
        errors.append("Please select a semester.")
    if not excel_file:
        errors.append("Please upload an Excel file.")
    elif not excel_file.name.endswith(".xlsx"):
        errors.append("Only .xlsx files are supported.")

    if errors:
        ctx["errors"] = errors
        return render(request, "results/import_marks.html", ctx)

    try:
        department = Department.objects.get(id=department_id)
        year = Year.objects.get(id=year_id)
    except (Department.DoesNotExist, Year.DoesNotExist):
        ctx["errors"] = ["Invalid department or year."]
        return render(request, "results/import_marks.html", ctx)

    semester_int = int(semester_id)
    subjects = list(Subject.objects.filter(department=department, semester=semester_int))

    if not subjects:
        ctx["errors"] = [
            f"❌ No subjects found for {department.name} - Semester {semester_int}. "
            "Please add Subjects in the Admin Panel before importing marks."
        ]
        return render(request, "results/import_marks.html", ctx)

    # ── Save the uploaded file to disk ─────────────────────────────────────
    import uuid
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile

    safe_name = f"{department_id}_{year_id}_{semester_id}_{uuid.uuid4().hex[:8]}_{excel_file.name}"
    save_path = os.path.join("uploads", "marks", safe_name)
    saved_file_path = default_storage.save(save_path, ContentFile(excel_file.read()))
    full_saved_path = os.path.join(settings.MEDIA_ROOT, saved_file_path)

    try:
        wb = openpyxl.load_workbook(full_saved_path)
        ws = wb.active

        # ── Validate header – marks file must have more than 3 columns ──
        header_row = [
            str(c).strip()
            for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if c is not None and str(c).strip() != ""
        ]
        if len(header_row) <= 3:
            wrong_cols = ", ".join(header_row[:5]) + ("..." if len(header_row) > 5 else "")
            ctx["errors"] = [
                f"❌ Wrong file uploaded! This looks like a Students Excel file because it has "
                f"only {len(header_row)} columns (Found: {wrong_cols}). "
                "The Marks import file must have: Reg No, Name, then Internal & External for each subject. "
                "Please upload the correct Marks Excel file or use the 'Download Sample' button."
            ]
            return render(request, "results/import_marks.html", ctx)

        # ── Collect roll numbers from the xlsx ─────────────────────────────
        xlsx_rows = []  # list of tuples: (roll_no, row_data, row_index)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            xlsx_rows.append((str(row[0]).strip(), row, i))

        xlsx_roll_numbers = {r[0] for r in xlsx_rows}

        # ── Fetch all students in DB for this dept+year ────────────────────
        db_students = {
            s.roll_no: s
            for s in Student.objects.filter(department=department, year=year)
        }
        db_roll_numbers = set(db_students.keys())

        # ── Cross-validation sets ──────────────────────────────────────────
        matched_rolls   = xlsx_roll_numbers & db_roll_numbers   # in both
        not_in_db       = xlsx_roll_numbers - db_roll_numbers   # xlsx has, DB missing
        not_in_xlsx     = db_roll_numbers - xlsx_roll_numbers   # DB has, xlsx missing

        # ── Import marks for matched students ─────────────────────────────
        updated_count = 0
        row_errors = []
        matched_names = []

        for roll_no, row, i in xlsx_rows:
            if roll_no in not_in_db:
                row_errors.append(
                    f"Row {i}: ❌ '{roll_no}' — student not found in database for "
                    f"{department.name} / Year {year}. Import this student first."
                )
                continue

            student = db_students[roll_no]
            col_index = 2
            subject_errors = []

            for subject in subjects:
                internal_raw = row[col_index] if col_index < len(row) else None
                external_raw = row[col_index + 1] if col_index + 1 < len(row) else None

                try:
                    internal = int(internal_raw) if internal_raw not in [None, ""] else 0
                except Exception:
                    internal = 0
                try:
                    external = int(external_raw) if external_raw not in [None, ""] else 0
                except Exception:
                    external = 0

                try:
                    Mark.objects.update_or_create(
                        student=student, subject=subject, semester=semester_int,
                        defaults={"internal": internal, "external": external}
                    )
                except ValueError as ve:
                    subject_errors.append(f"{subject.name}: {ve}")

                col_index += 2

            if subject_errors:
                row_errors.append(f"Row {i} ({roll_no}): " + "; ".join(subject_errors))
            else:
                updated_count += 1
                matched_names.append(f"{roll_no} – {student.name}")

        ctx.update({
            "success": True,
            "updated_count": updated_count,
            "row_errors": row_errors,
            "matched_rolls": sorted(matched_rolls),
            "not_in_db": sorted(not_in_db),
            "not_in_xlsx": sorted(not_in_xlsx),
            "department_id": department_id,
            "year_id": year_id,
            "saved_file": safe_name,
        })
        return render(request, "results/import_marks.html", ctx)

    except Exception as e:
        ctx["errors"] = [f"Failed to read Excel file: {str(e)}"]
        return render(request, "results/import_marks.html", ctx)


def download_marks_sample_excel(request):
    """Return a sample .xlsx for the Marks import format based on dept/semester subjects."""
    try:
        import openpyxl
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)

    department_id = request.GET.get("department")
    semester_id = request.GET.get("semester")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks"

    headers = ["Reg No", "Name"]
    if department_id and semester_id:
        subjects = Subject.objects.filter(department_id=department_id, semester=semester_id)
        for subj in subjects:
            headers.append(f"{subj.name} (Int)")
            headers.append(f"{subj.name} (Ext)")
    else:
        headers += ["Subject1 Int", "Subject1 Ext", "Subject2 Int", "Subject2 Ext"]

    ws.append(headers)
    ws.append(["21CS001", "John Doe"] + [0] * (len(headers) - 2))

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marks_sample.xlsx"'
    wb.save(response)
    return response